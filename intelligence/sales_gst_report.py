#!/usr/bin/env python3
"""
Taiwan Maami — Sales & GST Report (Daily / Weekly / Monthly)

Outputs:
  1. HTML summary report (all periods) — outlet x channel with revenue and GST
  2. Excel bill-level detail report (daily only) — one row per bill, all outlets

Usage:
  python intelligence/sales_gst_report.py --period daily
  python intelligence/sales_gst_report.py --period weekly
  python intelligence/sales_gst_report.py --period monthly

Required env vars:
  SUPABASE_URL, SUPABASE_SERVICE_KEY
"""

import argparse
import os
import sys
from collections import defaultdict, OrderedDict
from datetime import date, datetime, timedelta

import pytz
from supabase import create_client


# --- Date range helpers -------------------------------------------------------

def get_date_range(period: str) -> tuple[date, date, str]:
    ist = pytz.timezone("Asia/Kolkata")
    today = datetime.now(ist).date()

    if period == "daily":
        d = today - timedelta(days=1)
        return d, d, d.strftime("%-d %B %Y")

    elif period == "weekly":
        days_since_monday = today.weekday()
        last_monday = today - timedelta(days=days_since_monday + 7)
        last_sunday = last_monday + timedelta(days=6)
        label = f"Week of {last_monday.strftime('%-d')}-{last_sunday.strftime('%-d %B %Y')}"
        return last_monday, last_sunday, label

    elif period == "monthly":
        first_of_this_month = today.replace(day=1)
        last_day_prev = first_of_this_month - timedelta(days=1)
        first_day_prev = last_day_prev.replace(day=1)
        label = first_day_prev.strftime("%B %Y")
        return first_day_prev, last_day_prev, label

    else:
        raise ValueError(f"Unknown period: {period}")


# --- Data fetching ------------------------------------------------------------

def fetch_raw_rows(start_date: date, end_date: date) -> list[dict]:
    """Fetch all clean_sales rows for the date range with pagination."""
    url = os.environ["SUPABASE_URL"]
    key = os.environ["SUPABASE_SERVICE_KEY"]
    client = create_client(url, key)

    start_str = start_date.isoformat()
    end_str = end_date.isoformat()

    all_rows = []
    page_size = 1000
    offset = 0

    while True:
        response = (
            client.table("clean_sales")
            .select(
                "outlet, channel_label, source_order_id, order_date, "
                "item_name, item_quantity, item_total_rupees, "
                "item_sequence, order_tax_rupees, order_discount_rupees, "
                "order_total_rupees, payment_method, aggregator"
            )
            .gte("order_date", start_str)
            .lte("order_date", end_str)
            .order("outlet")
            .order("source_order_id")
            .order("item_sequence")
            .range(offset, offset + page_size - 1)
            .execute()
        )
        batch = response.data or []
        all_rows.extend(batch)
        if len(batch) < page_size:
            break
        offset += page_size

    print(f"  Fetched {len(all_rows):,} line items from clean_sales")
    return all_rows


def aggregate_for_summary(all_rows: list[dict]) -> list[dict]:
    """Aggregate raw rows to outlet x channel summary."""
    OUTLET_DISPLAY = {
        "palladium": "Palladium",
        "tnagar": "T.Nagar",
        "annanagar": "Anna Nagar",
    }

    groups = defaultdict(lambda: {"orders": set(), "revenue": 0.0, "gst_collected": 0.0})

    for row in all_rows:
        outlet = row.get("outlet") or "Unknown"
        channel = row.get("channel_label") or "Unknown"
        key_tuple = (outlet, channel)
        groups[key_tuple]["orders"].add(row.get("source_order_id"))
        groups[key_tuple]["revenue"] += float(row.get("item_total_rupees") or 0)
        if row.get("item_sequence") == 0:
            groups[key_tuple]["gst_collected"] += float(row.get("order_tax_rupees") or 0)

    result = []
    for (outlet, channel), agg in sorted(groups.items()):
        display_name = OUTLET_DISPLAY.get(outlet, outlet.title())
        channel_display = channel.title() if channel else "Unknown"
        result.append({
            "outlet": display_name,
            "channel_label": channel_display,
            "orders": len(agg["orders"]),
            "revenue": agg["revenue"],
            "gst_collected": agg["gst_collected"],
        })

    return result


def aggregate_for_bill_report(all_rows: list[dict]) -> list[dict]:
    """Aggregate raw rows to one row per bill."""
    OUTLET_DISPLAY = {
        "palladium": "Palladium Mall",
        "tnagar": "T.Nagar",
        "annanagar": "Anna Nagar",
    }

    bill_order = list(OrderedDict.fromkeys(r["source_order_id"] for r in all_rows))
    bills = {}

    for row in all_rows:
        oid = row["source_order_id"]
        if oid not in bills:
            bills[oid] = {
                "bill_no":     oid,
                "date":        row.get("order_date", ""),
                "outlet":      OUTLET_DISPLAY.get(row.get("outlet", ""), row.get("outlet", "")),
                "channel":     "In-store" if row.get("channel_label") == "instore" else "Delivery",
                "payment":     (row.get("payment_method") or "Unknown").upper(),
                "aggregator":  row.get("aggregator") or "-",
                "items":       [],
                "item_count":  0,
                "gross_items": 0.0,
                "discount":    None,
                "gst":         None,
                "bill_total":  None,
            }
        qty = int(row.get("item_quantity") or 1)
        name = row.get("item_name") or "Unknown"
        bills[oid]["items"].append(f"{qty}x {name}" if qty > 1 else name)
        bills[oid]["item_count"] += qty
        bills[oid]["gross_items"] += float(row.get("item_total_rupees") or 0)
        if row.get("item_sequence") == 0:
            bills[oid]["discount"]   = float(row.get("order_discount_rupees") or 0)
            bills[oid]["gst"]        = float(row.get("order_tax_rupees") or 0)
            bills[oid]["bill_total"] = float(row.get("order_total_rupees") or 0)

    return [bills[oid] for oid in bill_order]


# --- Formatting helpers -------------------------------------------------------

def fmt_indian(amount: float) -> str:
    if amount < 0:
        return "-" + fmt_indian(-amount)
    rounded = round(amount, 2)
    integer_part = int(rounded)
    decimal_part = f".{int(round((rounded - integer_part) * 100)):02d}"
    s = str(integer_part)
    if len(s) <= 3:
        return s + decimal_part
    result = s[-3:]
    s = s[:-3]
    while s:
        result = s[-2:] + "," + result
        s = s[:-2]
    return result + decimal_part


# --- HTML summary report ------------------------------------------------------

def build_html_report(rows: list[dict], period_label: str) -> str:
    outlets: dict[str, list[dict]] = {}
    for row in rows:
        outlets.setdefault(row["outlet"], []).append(row)

    table_rows = ""
    grand_orders = 0
    grand_revenue = 0.0
    grand_gst = 0.0

    for outlet, channel_rows in outlets.items():
        outlet_orders = 0
        outlet_revenue = 0.0
        outlet_gst = 0.0

        for r in channel_rows:
            orders = int(r["orders"])
            revenue = float(r["revenue"])
            gst = float(r["gst_collected"])
            outlet_orders += orders
            outlet_revenue += revenue
            outlet_gst += gst

            table_rows += f"""
            <tr>
                <td style="padding:8px 12px; border-bottom:1px solid #eee;">{outlet}</td>
                <td style="padding:8px 12px; border-bottom:1px solid #eee;">{r['channel_label']}</td>
                <td style="padding:8px 12px; border-bottom:1px solid #eee; text-align:right;">{orders:,}</td>
                <td style="padding:8px 12px; border-bottom:1px solid #eee; text-align:right;">{fmt_indian(revenue)}</td>
                <td style="padding:8px 12px; border-bottom:1px solid #eee; text-align:right;">{fmt_indian(gst)}</td>
            </tr>"""

        grand_orders += outlet_orders
        grand_revenue += outlet_revenue
        grand_gst += outlet_gst

        table_rows += f"""
        <tr style="background:#f8f8f8; font-weight:600;">
            <td style="padding:8px 12px; border-bottom:2px solid #ddd;">{outlet} Total</td>
            <td style="padding:8px 12px; border-bottom:2px solid #ddd;"></td>
            <td style="padding:8px 12px; border-bottom:2px solid #ddd; text-align:right;">{outlet_orders:,}</td>
            <td style="padding:8px 12px; border-bottom:2px solid #ddd; text-align:right;">{fmt_indian(outlet_revenue)}</td>
            <td style="padding:8px 12px; border-bottom:2px solid #ddd; text-align:right;">{fmt_indian(outlet_gst)}</td>
        </tr>"""

    table_rows += f"""
    <tr style="background:#2d2d2d; color:#fff; font-weight:700;">
        <td style="padding:10px 12px;">GRAND TOTAL</td>
        <td style="padding:10px 12px;"></td>
        <td style="padding:10px 12px; text-align:right;">{grand_orders:,}</td>
        <td style="padding:10px 12px; text-align:right;">{fmt_indian(grand_revenue)}</td>
        <td style="padding:10px 12px; text-align:right;">{fmt_indian(grand_gst)}</td>
    </tr>"""

    return f"""<!DOCTYPE html>
<html>
<body style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;color:#333;padding:20px;">
    <h2 style="margin-bottom:4px;">Taiwan Maami - {period_label} Sales &amp; GST Summary</h2>
    <p style="color:#666;margin-top:0;">Generated on {datetime.now().strftime('%d %B %Y, %I:%M %p')}</p>
    <table style="border-collapse:collapse;width:100%;max-width:700px;font-size:14px;">
        <thead>
            <tr style="background:#9e0b0f;color:#fff;">
                <th style="padding:10px 12px;text-align:left;">Outlet</th>
                <th style="padding:10px 12px;text-align:left;">Channel</th>
                <th style="padding:10px 12px;text-align:right;">Orders</th>
                <th style="padding:10px 12px;text-align:right;">Revenue (Rs)</th>
                <th style="padding:10px 12px;text-align:right;">GST Collected (Rs)</th>
            </tr>
        </thead>
        <tbody>{table_rows}</tbody>
    </table>
    <p style="color:#999;font-size:12px;margin-top:20px;">Automated report - Taiwan Maami Intelligence.</p>
</body>
</html>"""


# --- Excel bill-level report --------------------------------------------------

def build_excel_report(bill_rows: list[dict], period_label: str, output_path: str):
    """Build one-row-per-bill Excel report for daily period."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  openpyxl not available - skipping Excel report")
        return

    NAVY = "1F3864"
    RED = "9E0B0F"
    WHITE = "FFFFFF"
    MID = "DDDDDD"
    LGRY = "F5F5F5"
    OUTLET_COLORS = {
        "Palladium Mall": "D6E4F0",
        "Anna Nagar": "E2EFDA",
        "T.Nagar": "FFF2CC",
    }
    INR = "#,##0.00"

    def fill(h):
        return PatternFill("solid", fgColor=h)

    ts = Side(style="thin", color=MID)
    TB = Border(left=ts, right=ts, top=ts, bottom=ts)

    def c(ws, r, col, v=None, bold=False, fc=None, ftc="000000",
          align="center", nf=None, size=10, wrap=False):
        cl = ws.cell(row=r, column=col, value=v)
        cl.font = Font(bold=bold, color=ftc, name="Arial", size=size)
        if fc:
            cl.fill = fill(fc)
        cl.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        cl.border = TB
        if nf:
            cl.number_format = nf
        return cl

    wb = Workbook()
    ws = wb.active
    ws.title = f"Bills - {period_label}"

    col_widths = [6, 22, 12, 18, 14, 12, 55, 10, 14, 12, 12, 14, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:M1")
    cl = ws["A1"]
    cl.value = f"TAIWAN MAAMI - DAILY BILL REPORT  |  {period_label}  |  All Outlets"
    cl.font = Font(bold=True, name="Arial", size=12, color=WHITE)
    cl.fill = fill(RED)
    cl.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:M2")
    cl = ws["A2"]
    cl.value = (
        f"One row per bill  |  {len(bill_rows)} bills  |  "
        "Palladium (blue)  T.Nagar (yellow)  Anna Nagar (green)  |  Source: Supabase clean_sales"
    )
    cl.font = Font(italic=True, name="Arial", size=9, color="595959")
    cl.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 5

    hdrs = [
        "#", "Bill No.", "Date", "Outlet", "Mode of Sale", "Aggregator",
        "Items on Bill", "Item Count", "Gross Items Rs",
        "Discount Rs", "GST Rs", "Bill Total Rs", "Mode of Payment",
    ]
    ws.row_dimensions[4].height = 40
    for i, h in enumerate(hdrs, 1):
        c(ws, 4, i, h, bold=True, fc=NAVY, ftc=WHITE, wrap=True)

    r = 5
    for idx, b in enumerate(bill_rows, 1):
        ws.row_dimensions[r].height = 22
        rc = OUTLET_COLORS.get(b["outlet"], LGRY)
        items_str = " | ".join(b["items"])

        c(ws, r,  1, idx,              fc=rc, align="center", size=9)
        c(ws, r,  2, b["bill_no"],     fc=rc, align="left",   size=9)
        c(ws, r,  3, b["date"],        fc=rc, align="center", size=9)
        c(ws, r,  4, b["outlet"],      fc=rc, align="center", size=9)
        c(ws, r,  5, b["channel"],     fc=rc, align="center", size=9)
        c(ws, r,  6, b["aggregator"],  fc=rc, align="center", size=9)
        c(ws, r,  7, items_str,        fc=rc, align="left",   size=8, wrap=True)
        c(ws, r,  8, b["item_count"],  fc=rc, align="center", size=9)
        c(ws, r,  9, b["gross_items"], fc=rc, align="right",  nf=INR, size=9)
        c(ws, r, 10, b["discount"] if b["discount"] is not None else 0,
                                       fc=rc, align="right",  nf=INR, size=9)
        c(ws, r, 11, b["gst"] if b["gst"] is not None else 0,
                                       fc=rc, align="right",  nf=INR, size=9)
        c(ws, r, 12, b["bill_total"] if b["bill_total"] is not None else 0,
                                       fc=rc, align="right",  nf=INR, size=9)
        c(ws, r, 13, b["payment"],     fc=rc, align="center", size=9)
        r += 1

    ws.row_dimensions[r].height = 5
    r += 1
    ws.row_dimensions[r].height = 24

    t_bills  = len(bill_rows)
    t_items  = sum(b["item_count"] for b in bill_rows)
    t_gross  = sum(b["gross_items"] for b in bill_rows)
    t_disc   = sum(b["discount"] or 0 for b in bill_rows)
    t_gst    = sum(b["gst"] or 0 for b in bill_rows)
    t_total  = sum(b["bill_total"] or 0 for b in bill_rows)

    ws.merge_cells(f"A{r}:F{r}")
    cl = ws[f"A{r}"]
    cl.value = f"GRAND TOTAL - {t_bills} bills"
    cl.font = Font(bold=True, name="Arial", size=10, color=WHITE)
    cl.fill = fill(NAVY)
    cl.alignment = Alignment(horizontal="center", vertical="center")
    cl.border = TB

    for col, val in [
        (7, None), (8, t_items), (9, t_gross),
        (10, t_disc), (11, t_gst), (12, t_total), (13, None),
    ]:
        c(ws, r, col, val, bold=True, fc=NAVY, ftc=WHITE,
          align="right" if isinstance(val, (int, float)) else "center",
          nf=INR if isinstance(val, float) or (col > 8 and isinstance(val, int) and col != 8)
          else "#,##0" if col == 8 else None)

    r += 2
    ws.merge_cells(f"A{r}:M{r}")
    cl = ws[f"A{r}"]
    cl.value = "SUMMARY BY OUTLET"
    cl.font = Font(bold=True, name="Arial", size=10, color=WHITE)
    cl.fill = fill(NAVY)
    cl.alignment = Alignment(horizontal="center", vertical="center")
    cl.border = TB
    ws.row_dimensions[r].height = 20
    r += 1

    ws.row_dimensions[r].height = 18
    for i, h in enumerate(
        ["Outlet", "Bills", "Items", "Gross Items Rs", "Discount Rs", "GST Rs", "Bill Total Rs"], 1
    ):
        c(ws, r, i, h, bold=True, fc=NAVY, ftc=WHITE)
    r += 1

    ot = defaultdict(lambda: [0, 0, 0, 0, 0, 0])
    for b in bill_rows:
        o = b["outlet"]
        ot[o][0] += 1
        ot[o][1] += b["item_count"]
        ot[o][2] += b["gross_items"]
        ot[o][3] += b["discount"] or 0
        ot[o][4] += b["gst"] or 0
        ot[o][5] += b["bill_total"] or 0

    for ok in ["Palladium Mall", "Anna Nagar", "T.Nagar"]:
        if ok not in ot:
            continue
        v = ot[ok]
        rc = OUTLET_COLORS.get(ok, LGRY)
        ws.row_dimensions[r].height = 20
        for i, val in enumerate([ok, v[0], v[1], v[2], v[3], v[4], v[5]], 1):
            al = "left" if i == 1 else "right" if i > 2 else "center"
            nf = INR if i > 3 else ("#,##0" if i in (2, 3) else None)
            c(ws, r, i, val, fc=rc, align=al, nf=nf, bold=True)
        r += 1

    ws.row_dimensions[r].height = 22
    for i, val in enumerate(
        ["GRAND TOTAL", t_bills, t_items, t_gross, t_disc, t_gst, t_total], 1
    ):
        al = "left" if i == 1 else "right" if i > 2 else "center"
        nf = INR if i > 3 else ("#,##0" if i in (2, 3) else None)
        c(ws, r, i, val, bold=True, fc=NAVY, ftc=WHITE, align=al, nf=nf)
    r += 2

    ws.merge_cells(f"A{r}:M{r}")
    cl = ws[f"A{r}"]
    cl.value = "SETTLEMENT BY PAYMENT METHOD"
    cl.font = Font(bold=True, name="Arial", size=10, color=WHITE)
    cl.fill = fill(NAVY)
    cl.alignment = Alignment(horizontal="center", vertical="center")
    cl.border = TB
    ws.row_dimensions[r].height = 20
    r += 1

    ws.row_dimensions[r].height = 18
    for i, h in enumerate(["Payment Method", "Bills", "Bill Total Rs"], 1):
        c(ws, r, i, h, bold=True, fc=NAVY, ftc=WHITE)
    r += 1

    pmt = defaultdict(lambda: [0, 0.0])
    for b in bill_rows:
        pmt[b["payment"]][0] += 1
        pmt[b["payment"]][1] += b["bill_total"] or 0

    for pm, (cnt, amt) in sorted(pmt.items()):
        ws.row_dimensions[r].height = 18
        c(ws, r, 1, pm,  fc=LGRY, align="left",  bold=True)
        c(ws, r, 2, cnt, fc=LGRY, align="center", nf="#,##0")
        c(ws, r, 3, amt, fc=LGRY, align="right",  nf=INR)
        for col in range(4, 14):
            c(ws, r, col, fc=LGRY)
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:M{r}")
    cl = ws[f"A{r}"]
    cl.value = "SUMMARY BY CHANNEL"
    cl.font = Font(bold=True, name="Arial", size=10, color=WHITE)
    cl.fill = fill(NAVY)
    cl.alignment = Alignment(horizontal="center", vertical="center")
    cl.border = TB
    ws.row_dimensions[r].height = 20
    r += 1

    ws.row_dimensions[r].height = 18
    for i, h in enumerate(["Channel", "Bills", "Bill Total Rs"], 1):
        c(ws, r, i, h, bold=True, fc=NAVY, ftc=WHITE)
    r += 1

    ch_t = defaultdict(lambda: [0, 0.0])
    for b in bill_rows:
        ch_t[b["channel"]][0] += 1
        ch_t[b["channel"]][1] += b["bill_total"] or 0

    for ch, (cnt, amt) in sorted(ch_t.items()):
        ws.row_dimensions[r].height = 18
        c(ws, r, 1, ch,  fc=LGRY, align="left",  bold=True)
        c(ws, r, 2, cnt, fc=LGRY, align="center", nf="#,##0")
        c(ws, r, 3, amt, fc=LGRY, align="right",  nf=INR)
        for col in range(4, 14):
            c(ws, r, col, fc=LGRY)
        r += 1

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"  Excel bill report saved: {output_path}")
    print(f"    {t_bills} bills | {t_items} items | Rs {fmt_indian(t_total)} total collected")


# --- Main ---------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Taiwan Maami Sales & GST Report")
    parser.add_argument("--period", choices=["daily", "weekly", "monthly"], required=True)
    args = parser.parse_args()

    ist = pytz.timezone("Asia/Kolkata")
    today_ist = datetime.now(ist).date()
    if args.period == "monthly" and today_ist.day != 1:
        print("Monthly report: today is not the 1st - skipping.")
        sys.exit(0)

    start_date, end_date, period_label = get_date_range(args.period)
    print(f"Period: {args.period} | Range: {start_date} to {end_date} | Label: {period_label}")

    all_rows = fetch_raw_rows(start_date, end_date)

    os.makedirs("reports", exist_ok=True)

    summary_rows = aggregate_for_summary(all_rows)
    html = build_html_report(summary_rows, period_label)
    html_path = f"reports/sales_gst_{args.period}_{start_date.isoformat()}.html"
    with open(html_path, "w") as f:
        f.write(html)
    print(f"  HTML summary saved: {html_path}")

    if args.period == "daily":
        bill_rows = aggregate_for_bill_report(all_rows)
        xlsx_path = f"reports/daily_bills_{start_date.isoformat()}.xlsx"
        build_excel_report(bill_rows, period_label, xlsx_path)


if __name__ == "__main__":
    main()
